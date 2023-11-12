import xlwt
from openpyxl import load_workbook
from utils import get_client_data, get_product_info
from datetime import datetime
import locale

def create_xls_file(output, client, data, catalog_number, quantity):
    '''
    Creates an Excel file

    Args:
    output: list[float] - a list with profitability values ​​predicted by the model 
    client: str - the client for whom the file will be created
    data: datetime - file creation date
    catalog_number: list[str] - list of catalog numbers that will be offered to the client
    quantity: int - the quantity of the product 

    Returns Excel-file

    Note: The Excel template file used to create a new file is located in the "cp" folder with the name "sample.xlsx".

    Example usage: create_xls_file(output=[30.124], client='Client', data=datetime.now, catalog_number=['SPR00SIA1'], quantity=1)

    '''

    locale.setlocale(locale.LC_TIME, "ru_RU")
    template_filename = "./cp/sample.xlsx"
    workbook = load_workbook(template_filename)

    # Выбираем нужный лист в шаблоне
    sheet = workbook["Лист1"]

    cell = sheet.cell(row=9, column=2)
    cell.value = get_client_data(client, 'full_name')

    cell = sheet.cell(row=11, column=4)
    cell.value = data.strftime("%d %B %Y")

        # Заполняем необходимые ячейки из вывода модели
    for i, value in enumerate(output):
        cell = sheet.cell(row=i+18, column=2)
        cell.value = i+1

        cell = sheet.cell(row=i+18, column=3)
        cell.value = catalog_number[i]

        cell = sheet.cell(row=i+18, column=4)
        cell.value = get_product_info(catalog_number[i], 'name')

        cell = sheet.cell(row=i+18, column=5)
        cell.value = quantity

        cell = sheet.cell(row=i+18, column=6)
        cell.value = round((get_product_info(catalog_number[i], 'price') * (1 + ((get_product_info(catalog_number[i], 'transport') + get_product_info(catalog_number[i], 'CD'))/100)) * (100 + output[i])/100), 2)

        cell = sheet.cell(row=i+18, column=7)
        cell.value = sheet.cell(row=i+18, column=5).value * sheet.cell(row=i+18, column=6).value

        cell = sheet.cell(row=i+18, column=8)
        cell.value = round(sheet.cell(row=i+18, column=7).value * 0.2, 2)

        cell = sheet.cell(row=i+18, column=9)
        cell.value = round(sheet.cell(row=i+18, column=7).value * 1.2, 2)

    return workbook